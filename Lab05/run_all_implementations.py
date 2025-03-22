import os
import subprocess
import re
import statistics
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

def run_nbody_program(paths, arguments, output_excel):
    # Create an Excel workbook and sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "NBody Results"

    # Write the headers with styling
    headers = ["Path", "Bodies", "Average Time (s)", "Std Dev Time (s)", "Average Interactions (Billion/s)"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Track current row in Excel
    current_row = 2

    # Iterate over each path and argument
    for path in paths:
        code_dir = os.path.join(path, "Code")
        program_path = os.path.join(code_dir, "nbody")

        # Ensure the program is built
        if not os.path.exists(program_path):
            try:
                print(f"Building executable in {code_dir}")
                subprocess.run(["make", "clean"], cwd=code_dir, check=True)
                subprocess.run(["make"], cwd=code_dir, check=True)
            except subprocess.CalledProcessError as e:
                print(f"Error building program in {code_dir}: {e}")
                continue

        for arg in arguments:
            print(f"Running {program_path} with argument {arg}")
            try:
                # Run the program and capture the output
                output = subprocess.run([program_path, str(arg)], capture_output=True, text=True, check=True).stdout

                # Extract iteration times and interaction rate
                iteration_times = []
                avg_interactions = None
                for line in output.splitlines():
                    # print(f"Processing line: {line}")  # Debug log for each line
                    match_time = re.match(r"Iteration \d+: ([\d\.eE\+\-]+) seconds", line)
                    match_interactions = re.match(rf"{arg} Bodies: average ([\d\.]+) Billion Interactions / second", line)

                    if match_time:
                        # print(f"Matched iteration time: {match_time.group(1)}")  # Debug the matched time
                        iteration_times.append(float(match_time.group(1)))
                    elif match_interactions:
                        # print(f"Matched interactions: {match_interactions.group(1)}")  # Debug the matched interactions
                        avg_interactions = float(match_interactions.group(1))

                # Debugging logs
                print(f"Iteration times for {arg}: {iteration_times}")
                print(f"Average interactions for {arg}: {avg_interactions}")

                if len(iteration_times) > 1:
                    iteration_times = iteration_times[1:]  # Skip the first iteration
                    avg_time = statistics.mean(iteration_times)
                    std_dev_time = statistics.stdev(iteration_times)

                    # Format runtime in scientific notation
                    avg_time_sci = f"{avg_time:.3e}"
                    std_dev_time_sci = f"{std_dev_time:.3e}"

                    # Write results to the Excel file
                    sheet.cell(row=current_row, column=1, value=path)
                    sheet.cell(row=current_row, column=2, value=arg)
                    sheet.cell(row=current_row, column=3, value=avg_time_sci)
                    sheet.cell(row=current_row, column=4, value=std_dev_time_sci)
                    sheet.cell(row=current_row, column=5, value=avg_interactions)

                    # Apply alignment and styling for this row
                    for col in range(1, 6):
                        cell = sheet.cell(row=current_row, column=col)
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    # Log current row for debugging
                    # print(f"Data written to row {current_row} for argument {arg}")
                    current_row += 1
                else:
                    print(f"Not enough data to calculate results for argument {arg}")

            except subprocess.CalledProcessError as e:
                print(f"Error running program at {program_path} with argument {arg}: {e}")
            except Exception as e:
                print(f"Unexpected error: {e}")

    # Auto-adjust column widths
    for column_cells in sheet.columns:
        max_length = 0
        col_letter = column_cells[0].column_letter  # Get the column letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[col_letter].width = adjusted_width

    # Save the Excel workbook
    workbook.save(output_excel)
    print(f"Results saved to {output_excel}")


paths = [
            # "/srv/homes/gkapakos/Lab05/CPU_Original"
            "/srv/homes/gkapakos/Lab05/OMP_Accelerated",
            "/srv/homes/gkapakos/Lab05/OMP_Accelerated_2",
            "/srv/homes/gkapakos/Lab05/CUDA_0_no_optimization",
            "/srv/homes/gkapakos/Lab05/CUDA_1_registers",
            "/srv/homes/gkapakos/Lab05/CUDA_2_SoA",
            "/srv/homes/gkapakos/Lab05/CUDA_3_SoA_float3",
            "/srv/homes/gkapakos/Lab05/CUDA_4_SoA_float4",
            "/srv/homes/gkapakos/Lab05/CUDA_5_tiling_shared",
            "/srv/homes/gkapakos/Lab05/CUDA_6_unroll",
            "/srv/homes/gkapakos/Lab05/CUDA_7_fast_math",
            "/srv/homes/gkapakos/Lab05/CUDA_8_final"
        ]

arguments = [32768, 65536, 131072, 262144]
output_excel = "Runtime_results_v2.xlsx"

run_nbody_program(paths, arguments, output_excel)
