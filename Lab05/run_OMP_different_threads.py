import os
import subprocess
import re
import statistics
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

def run_nbody_program(paths, arguments, thread_counts, output_excel):
    # Create an Excel workbook and sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "NBody Results"

    # Write the headers with styling
    headers = ["Path", "Threads", "Bodies", "Average Time (s)", "Std Dev Time (s)", "Average Interactions (Billion/s)"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row = 2

    for path in paths:
        code_dir = os.path.join(path, "Code")
        program_path = os.path.join(code_dir, "nbody")

        # Ensure the program is built
        if not os.path.exists(program_path):
            try:
                print(f"Building executable in {code_dir}")
                subprocess.run(["make", "clean"], cwd=code_dir, check=True)
                subprocess.run(["make"], cwd=code_dir, check=True)
            except subprocess.CalledProcessError as e:
                print(f"Error building program in {code_dir}: {e}")
                continue

        for threads in thread_counts:
            os.environ["OMP_NUM_THREADS"] = str(threads)
            print(f"Set OMP_NUM_THREADS={threads}")

            for arg in arguments:
                print(f"Running {program_path} with argument={arg} and threads={threads}")
                try:
                    output = subprocess.run(
                        [program_path, str(arg)],
                        capture_output=True,
                        text=True,
                        timeout=300,
                        check=True
                    ).stdout

                    print(f"Output:\n{output}")  # Debug subprocess output

                    # Extract data
                    iteration_times = []
                    avg_interactions = None
                    for line in output.splitlines():
                        match_time = re.match(r"Iteration \d+: ([\d\.eE\+\-]+) seconds", line)
                        match_interactions = re.match(rf"{arg} Bodies: average ([\d\.]+) Billion Interactions / second", line)
                        print(f"Line: {line}, Match Time: {match_time}, Match Interactions: {match_interactions}")

                        if match_time:
                            iteration_times.append(float(match_time.group(1)))
                        elif match_interactions:
                            avg_interactions = float(match_interactions.group(1))

                    if len(iteration_times) > 1:
                        iteration_times = iteration_times[1:]  # Skip first iteration
                        avg_time = statistics.mean(iteration_times)
                        std_dev_time = statistics.stdev(iteration_times)

                        avg_time_sci = f"{avg_time:.3e}"
                        std_dev_time_sci = f"{std_dev_time:.3e}"

                        # Write to Excel
                        print(f"Writing results for path={path}, threads={threads}, bodies={arg}")
                        sheet.cell(row=current_row, column=1, value=path)
                        sheet.cell(row=current_row, column=2, value=threads)
                        sheet.cell(row=current_row, column=3, value=arg)
                        sheet.cell(row=current_row, column=4, value=avg_time_sci)
                        sheet.cell(row=current_row, column=5, value=std_dev_time_sci)
                        sheet.cell(row=current_row, column=6, value=avg_interactions)

                        for col in range(1, 7):
                            sheet.cell(row=current_row, column=col).alignment = Alignment(horizontal="center", vertical="center")

                        current_row += 1
                    else:
                        print(f"Not enough data for threads={threads}, bodies={arg}")

                except subprocess.TimeoutExpired:
                    print(f"Program timed out for argument={arg} with threads={threads}")
                except subprocess.CalledProcessError as e:
                    print(f"Error running program at {program_path} with argument={arg} and threads={threads}: {e}")
                except Exception as e:
                    print(f"Unexpected error: {e}")

    # Save Excel
    workbook.save(output_excel)
    print(f"Results saved to {output_excel}")

paths = [
            "/srv/homes/gkapakos/Lab05/OMP_Accelerated",
            "/srv/homes/gkapakos/Lab05/OMP_Accelerated_2"
        ]
arguments = [131072]
thread_counts = [4, 8, 14, 28, 56]
# thread_counts = [16]
output_excel = "Runtime_results_threads.xlsx"

run_nbody_program(paths, arguments, thread_counts, output_excel)
