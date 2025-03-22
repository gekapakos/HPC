import os
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import random


def generate_colors(implementations):
    """Generate consistent colors for each implementation."""
    color_map = {}
    for implementation in implementations:
        if implementation not in color_map:
            color_map[implementation] = tuple(random.random() for _ in range(3))
    return color_map


def plot_gpu_time_with_stddev_from_excel(excel_path, output_plot_dir):
    # Load the Excel workbook
    wb = load_workbook(excel_path)
    os.makedirs(output_plot_dir, exist_ok=True)

    # Create a global color map for all implementations
    all_implementations = set()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            all_implementations.add(row[0])  # Collect all unique implementations
    color_map = generate_colors(all_implementations)

    # Iterate through each sheet (representing images)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Read data from the sheet
        implementations = []
        gpu_times = []
        std_devs = []

        # Skip the header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            implementations.append(row[0])  # Executable Directory
            gpu_times.append(row[1])  # GPU Time (s)
            std_devs.append(row[2])  # Standard Deviation GPU (s)

        # Plotting
        plt.figure(figsize=(10, 6))
        bar_colors = [color_map[impl] for impl in implementations]
        bars = plt.bar(implementations, gpu_times, color=bar_colors, label="GPU Time (s)")
        
        # Overlay line for standard deviation
        plt.plot(implementations, std_devs, color="red", marker="o", label="Standard Deviation (s)", linewidth=2)

        plt.xlabel("Implementation")
        plt.ylabel("Time (s)")
        plt.title(f"GPU Execution Time with Standard Deviation for {sheet_name}")
        plt.xticks(rotation=45, ha="right")
        plt.tight_layout()

        # Annotate bars with GPU time values in 0.00000 format
        for bar, time in zip(bars, gpu_times):
            plt.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height(),
                f"{time:.5f}",
                ha="center",
                va="bottom",
                fontsize=10
            )

        # Annotate the standard deviation line with values in 0.00000 format
        for i, std in enumerate(std_devs):
            plt.text(
                i,
                std,
                f"{std:.5f}",
                ha="center",
                va="bottom",
                color="red",
                fontsize=10
            )

        plt.legend()
        
        # Save plot to output directory
        plot_file_path = os.path.join(output_plot_dir, f"{sheet_name}_gpu_time_with_stddev.png")
        plt.savefig(plot_file_path)
        plt.close()

        print(f"Plot saved: {plot_file_path}")


# Example usage
if __name__ == "__main__":
    excel_path = "/srv/homes/gkapakos/Lab04/output_results_v3.xlsx"
    output_plot_dir = "/srv/homes/gkapakos/Lab04/plots_standard_deviation"

    plot_gpu_time_with_stddev_from_excel(excel_path, output_plot_dir)
