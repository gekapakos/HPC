import os
import subprocess
import filecmp
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import statistics


def run_contrast_enhancement(executable_dirs, image_dir, correct_output_dir, excel_output_path, iterations=1):
    # Prepare to collect results
    image_results = {}

    for exec_dir in executable_dirs:
        executable_path = os.path.join(exec_dir, "Code", "contrast-enhancement")
        output_dir = os.path.join(exec_dir, "..", "Output_Images")
        os.makedirs(output_dir, exist_ok=True)

        # Verify the executable exists
        if not os.path.isfile(executable_path):
            print(f"Executable not found: {executable_path}")
            continue

        print(f"Running executable: {executable_path}")

        for image_name in os.listdir(image_dir):
            input_image_path = os.path.join(image_dir, image_name)

            # Skip non-image files
            if not os.path.isfile(input_image_path):
                continue

            # Initialize results for the image if not already
            if image_name not in image_results:
                image_results[image_name] = []

            # Define output image path
            output_image_name = f"{os.path.splitext(image_name)[0]}_out.png"
            output_image_path = os.path.join(output_dir, output_image_name)

            # Store execution times for multiple iterations
            gpu_times = []
            cpu_times = []

            for i in range(iterations):
                try:
                    # Construct the command
                    command = [
                        executable_path,
                        input_image_path,
                        output_image_path
                    ]

                    # Run the command and capture stdout
                    process = subprocess.run(
                        command,
                        capture_output=True,
                        text=True,
                        check=True
                    )

                    # Extract execution times from stdout
                    stdout = process.stdout
                    for line in stdout.splitlines():
                        if "GPU Execution time:" in line:
                            gpu_times.append(float(line.split(":")[-1].strip().split()[0]))
                        elif "CPU Execution time:" in line:
                            cpu_times.append(float(line.split(":")[-1].strip().split()[0]))

                except subprocess.CalledProcessError as e:
                    print(f"Error running command {command}: {e.stderr}")

            # Remove best and worst times (outliers)
            if len(gpu_times) > 2:
                gpu_times.sort()
                cpu_times.sort()
                gpu_times = gpu_times[1:-1]
                cpu_times = cpu_times[1:-1]

            # Calculate average and standard deviation
            avg_gpu_time = statistics.mean(gpu_times) if gpu_times else None
            std_gpu_time = statistics.stdev(gpu_times) if len(gpu_times) > 1 else None

            avg_cpu_time = statistics.mean(cpu_times) if cpu_times else None
            std_cpu_time = statistics.stdev(cpu_times) if len(cpu_times) > 1 else None

            # Compare output image with correct output
            correct_output_path = os.path.join(correct_output_dir, output_image_name)
            is_correct = False
            if os.path.isfile(correct_output_path):
                is_correct = filecmp.cmp(output_image_path, correct_output_path, shallow=False)

            # Save the results for the image
            if avg_gpu_time is not None and avg_cpu_time is not None:
                image_results[image_name].append({
                    "Executable Directory": exec_dir,
                    "Average GPU Time (s)": avg_gpu_time,
                    "Standard Deviation GPU (s)": std_gpu_time,
                    "Average CPU Time (s)": avg_cpu_time,
                    "Standard Deviation CPU (s)": std_cpu_time,
                    "Is Correct": "Yes" if is_correct else "No"
                })

    # Write results to an Excel file
    write_results_to_excel(image_results, excel_output_path)

    print(f"Results saved to {excel_output_path}")


def write_results_to_excel(image_results, excel_path):
    wb = Workbook()
    for image_name, results in image_results.items():
        ws = wb.create_sheet(title=os.path.splitext(image_name)[0])

        # Define headers
        headers = [
            "Executable Directory",
            "Average GPU Time (s)",
            "Standard Deviation GPU (s)",
            "Average CPU Time (s)",
            "Standard Deviation CPU (s)",
            "Is Correct"
        ]
        ws.append(headers)

        # Apply styles to headers
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Write data for the image
        for result in results:
            ws.append([
                result["Executable Directory"],
                result["Average GPU Time (s)"],
                result["Standard Deviation GPU (s)"],
                result["Average CPU Time (s)"],
                result["Standard Deviation CPU (s)"],
                result["Is Correct"]
            ])

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)  # Get column letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Save the workbook
    wb.save(excel_path)



# Example usage
if __name__ == "__main__":
    executable_dirs = [
        # "CUDA_0_No_optimisation/",
        # "CUDA_1_shared_memory_histogram/",
        # "CUDA_2_shared_memory_loop/",
        # "CUDA_3_shared_memory_histogram_equalization/",
        # "CUDA_4_pinned_memory/",
        # "CUDA_5_pinned_memory_mapped/",
        # "CUDA_6_streams/",
        # "CUDA_7_unified_memory/",
        # "CUDA_8_constant_memory/",
        # "CUDA_9_texture_memory/",
        "CUDA_Final_all_optimisations/"
        # "CUDA_fix_streams/"

        # "CUDA_0/",
        # "CUDA_1/",
        # "CUDA_2/",
        # "CUDA_3/",
        # "CUDA_4/",
        # "CUDA_5/",
        # "CUDA_5_1/",
        # "CUDA_6/",
        # "CUDA_7/",
        # "CUDA_8/",
        # "CUDA_9/",
        # "CUDA_10/",
        # "CUDA_11/",
        # "CUDA_filippou/",
        # "CUDA_Final/"
    ]
    image_dir = "/srv/homes/gkapakos/Lab04/Original/Images"
    correct_output_dir = "/srv/homes/gkapakos/Lab04/Original/Output_Images"
    excel_output_path = "/srv/homes/gkapakos/Lab04/output_results_v4.xlsx"

    run_contrast_enhancement(executable_dirs, image_dir, correct_output_dir, excel_output_path)
