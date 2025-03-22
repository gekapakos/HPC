import os
from openpyxl import load_workbook
import matplotlib.pyplot as plt


def plot_average_cpu_time_per_image(excel_path, output_plot_dir):
    # Load the Excel workbook
    wb = load_workbook(excel_path)
    os.makedirs(output_plot_dir, exist_ok=True)

    # Target implementation
    target_implementation = "CPU_implementation"

    # Prepare data for plotting
    sheet_names = wb.sheetnames
    images = []
    avg_cpu_times = []

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        total_cpu_time = 0
        count = 0

        # Collect data from the sheet
        for row in ws.iter_rows(min_row=2, values_only=True):
            implementation = row[0]  # Executable Directory
            cpu_time = row[1]  # CPU Time (s)

            if implementation == target_implementation:
                total_cpu_time += cpu_time
                count += 1

        # Calculate average CPU time for the sheet (image)
        if count > 0:
            images.append(sheet_name)
            avg_cpu_times.append(total_cpu_time / count)

    # Plotting
    plt.figure(figsize=(10, 6))
    bars = plt.bar(images, avg_cpu_times, color='orange')
    plt.xlabel("Image")
    plt.ylabel("Average CPU Time (s)")
    plt.title(f"Average CPU Execution Time for {target_implementation}")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()

    # Annotate bars with average CPU times in 0.00000 format
    for bar, avg_time in zip(bars, avg_cpu_times):
        plt.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height(),
            f"{avg_time:.5f}",
            ha="center",
            va="bottom",
            fontsize=10
        )

    # Save the plot
    plot_file_path = os.path.join(output_plot_dir, f"{target_implementation}_average_cpu_time_per_image.png")
    plt.savefig(plot_file_path)
    plt.close()

    print(f"Plot saved: {plot_file_path}")


# Example usage
if __name__ == "__main__":
    excel_path = "/srv/homes/gkapakos/Lab04/output_results_v3.xlsx"
    output_plot_dir = "/srv/homes/gkapakos/Lab04/plots_avg_cpu_time"

    plot_average_cpu_time_per_image(excel_path, output_plot_dir)
