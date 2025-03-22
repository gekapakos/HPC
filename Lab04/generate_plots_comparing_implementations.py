import os
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import random


def plot_gpu_time_for_implementation_pairs(excel_path, output_plot_dir):
    # Load the Excel workbook
    wb = load_workbook(excel_path)
    os.makedirs(output_plot_dir, exist_ok=True)

    # Desired order of sheets to process
    sheet_order = ["fort", "uth", "x_ray", "ship", "planet_surface", "senator"]

    # Filter and order sheets
    available_sheets = [sheet for sheet in sheet_order if sheet in wb.sheetnames]

    # Create a mapping of implementation to image data
    implementation_data = {}

    for sheet_name in available_sheets:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            implementation = row[0]  # Executable Directory
            gpu_time = row[1]  # GPU Time (s)
            image_name = sheet_name

            if implementation not in implementation_data:
                implementation_data[implementation] = []
            implementation_data[implementation].append((image_name, gpu_time))

    # Get all implementations for comparison
    implementations = list(implementation_data.keys())

    # Generate a plot for each pair of consecutive implementations
    for i in range(len(implementations) - 1):  # Compare each implementation with the next one
        impl_1 = implementations[i]
        impl_2 = implementations[i + 1]

        # Get the data for the two implementations
        data_1 = implementation_data[impl_1]
        data_2 = implementation_data[impl_2]

        # Sort the data by the custom sheet order
        data_1.sort(key=lambda x: sheet_order.index(x[0]) if x[0] in sheet_order else float('inf'))
        data_2.sort(key=lambda x: sheet_order.index(x[0]) if x[0] in sheet_order else float('inf'))

        # Extract images and GPU times for both implementations
        images_1, gpu_times_1 = zip(*data_1)
        images_2, gpu_times_2 = zip(*data_2)

        # Make sure the images align (same sheet order)
        if images_1 != images_2:
            raise ValueError(f"Image order mismatch between {impl_1} and {impl_2}!")

        # Plotting
        plt.figure(figsize=(10, 6))

        # Bar plot for GPU time comparison
        bar_width = 0.35  # Width of the bars
        x = range(len(images_1))
        plt.bar(x, gpu_times_1, width=bar_width, label=impl_1, color='b', align='center')
        plt.bar([i + bar_width for i in x], gpu_times_2, width=bar_width, label=impl_2, color='g', align='center')

        plt.xlabel("Image")
        plt.ylabel("GPU Time (s)")
        plt.title(f"GPU Time Comparison: {impl_1} vs {impl_2}")
        plt.xticks([i + bar_width / 2 for i in x], images_1, rotation=45, ha="right")

        # Annotate bars with GPU time values in 0.00000 format
        for bar, time in zip(plt.gca().containers[0], gpu_times_1):
            plt.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height(),
                f"{time:.5f}",
                ha="center",
                va="bottom",
                fontsize=10
            )
        
        for bar, time in zip(plt.gca().containers[1], gpu_times_2):
            plt.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height(),
                f"{time:.5f}",
                ha="center",
                va="bottom",
                fontsize=10
            )

        # Add legends
        plt.legend(loc="upper left")

        # Tighten layout and save plot
        plt.tight_layout()

        # Save plot to output directory
        plot_file_path = os.path.join(output_plot_dir, f"{impl_1}_vs_{impl_2}_gpu_time_comparison.png")
        plt.savefig(plot_file_path)
        plt.close()

        print(f"Plot saved: {plot_file_path}")


# Example usage
if __name__ == "__main__":
    excel_path = "/srv/homes/gkapakos/Lab04/output_results_v3.xlsx"
    output_plot_dir = "/srv/homes/gkapakos/Lab04/plots_per_implementation_pair"

    plot_gpu_time_for_implementation_pairs(excel_path, output_plot_dir)
