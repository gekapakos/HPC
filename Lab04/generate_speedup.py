import os
from openpyxl import load_workbook
import matplotlib.pyplot as plt


def plot_speedup_for_senator(excel_path, output_plot_dir):
    # Load the Excel workbook
    wb = load_workbook(excel_path)
    os.makedirs(output_plot_dir, exist_ok=True)

    # Ensure "senator" sheet exists
    sheet_name = "senator"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"The sheet '{sheet_name}' is not found in the workbook.")

    # Access the "senator" sheet
    ws = wb[sheet_name]

    # Extract data
    implementations = []
    gpu_times = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        implementations.append(row[0])  # Executable Directory
        gpu_times.append(row[1])  # GPU Time (s)

    # Calculate speedup compared to CPU (first implementation)
    cpu_time = gpu_times[0]
    speedups = [cpu_time / time if time > 0 else 0 for time in gpu_times]

    # Plotting
    plt.figure(figsize=(10, 6))
    plt.plot(implementations, speedups, marker='o', linestyle='-', color='b', label="Speedup vs CPU")
    plt.xlabel("Implementation")
    plt.ylabel("Speedup (CPU Time / GPU Time)")
    plt.title("Speedup for 'senator' Image Across Implementations")
    plt.xticks(rotation=45, ha="right")
    plt.grid(True, linestyle="--", alpha=0.7)
    plt.legend(loc="upper left")

    # Annotate speedup values with "x<value>" format, shifted upward, and in orange
    for i, speedup in enumerate(speedups):
        plt.text(
            i,
            speedup + 0.1,  # Shift upward by 0.1 units
            f"x{speedup:.2f}",
            ha="center",
            va="bottom",
            fontsize=10,
            color="orange"
        )

    # Save plot to output directory
    plot_file_path = os.path.join(output_plot_dir, f"senator_speedup_vs_cpu.png")
    plt.tight_layout()
    plt.savefig(plot_file_path)
    plt.close()

    print(f"Plot saved: {plot_file_path}")


# Example usage
if __name__ == "__main__":
    excel_path = "/srv/homes/gkapakos/Lab04/output_results_v3.xlsx"
    output_plot_dir = "/srv/homes/gkapakos/Lab04/senator_speedup_plot"

    plot_speedup_for_senator(excel_path, output_plot_dir)
