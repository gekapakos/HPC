import os
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import random
import numpy as np


def plot_gpu_time_per_implementation_with_order_and_stddev(excel_path, output_plot_dir):
    # Load the Excel workbook
    wb = load_workbook(excel_path)
    os.makedirs(output_plot_dir, exist_ok=True)

    # Desired order of sheets to process
    sheet_order = ["fort", "uth", "x_ray", "ship", "planet_surface", "senator"]

    # Filter and order sheets
    available_sheets = [sheet for sheet in sheet_order if sheet in wb.sheetnames]

    # Create a mapping of implementation to image data
    implementation_data = {}

    for sheet_name in available_sheets:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            implementation = row[0]  # Executable Directory
            gpu_time = row[1]  # GPU Time (s)
            image_name = sheet_name
            stddev = row[2]  # Standard Deviation (s)

            if implementation not in implementation_data:
                implementation_data[implementation] = []
            implementation_data[implementation].append((image_name, gpu_time, stddev))

    # Generate a plot for each implementation
    for implementation, data in implementation_data.items():
        # Sort the data by the custom sheet order
        data.sort(key=lambda x: sheet_order.index(x[0]) if x[0] in sheet_order else float('inf'))
        images, gpu_times, std_devs = zip(*data)

        # Generate a random color for the bars
        bar_color = tuple(random.random() for _ in range(3))

        # Plotting
        fig, ax1 = plt.subplots(figsize=(10, 6))

        # Plot the bars for GPU time
        bars = ax1.bar(images, gpu_times, color=bar_color, label="GPU Time (s)")
        ax1.set_xlabel("Image")
        ax1.set_ylabel("GPU Time (s)")
        ax1.set_xticklabels(images, rotation=45, ha="right")
        
        # Annotate bars with GPU time values in 0.00000 format
        for bar, time in zip(bars, gpu_times):
            ax1.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height(),
                f"{time:.5f}",
                ha="center",
                va="bottom",
                fontsize=10
            )

        # Create a second y-axis for the standard deviation
        ax2 = ax1.twinx()
        ax2.plot(images, std_devs, color="blue", marker="o", label="Standard Deviation (s)", linewidth=2)
        ax2.set_ylabel("Standard Deviation (s)")

        # Annotate the standard deviation line with values in 0.00000 format, shifting the position a bit
        for i, std in enumerate(std_devs):
            ax2.text(
                i,
                std + 0.001,  # Shift the annotation slightly above the line
                f"{std:.5f}",
                ha="center",
                va="bottom",
                color="blue",
                fontsize=10
            )

        # Set title and layout
        plt.title(f"GPU Execution Time and Standard Deviation for {implementation}")
        plt.tight_layout()

        # Add legends for both axes
        ax1.legend(loc="upper left")
        ax2.legend(loc="upper right")

        # Save plot to output directory
        plot_file_path = os.path.join(output_plot_dir, f"{implementation}_gpu_time_with_stddev.png")
        plt.savefig(plot_file_path)
        plt.close()

        print(f"Plot saved: {plot_file_path}")


# Example usage
if __name__ == "__main__":
    excel_path = "/srv/homes/gkapakos/Lab04/output_results_v3.xlsx"
    output_plot_dir = "/srv/homes/gkapakos/Lab04/plots_per_implementation_with_stddev"

    plot_gpu_time_per_implementation_with_order_and_stddev(excel_path, output_plot_dir)
